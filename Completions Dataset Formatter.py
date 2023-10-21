import os
import sys
import csv
import json
import openpyxl

def read_from_excel(input_file_path):
    dataframe = openpyxl.load_workbook(input_file_path)
    dataframe1 = dataframe.active
    data=[]
    for row in range(0, dataframe1.max_row):
        rowData=[]
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            rowData.append(col[row].value)
        data.append(rowData)
    return data

def read_from_csv(input_file_path):
    with open(input_file_path, 'r', encoding='utf-8') as file:
        reader = csv.reader(file)
        data=[]
        for row in reader:
            data.append(row)
        return data
    
def write_to_jsonl(data, output_file_path):
    try:
        with open(output_file_path, 'w', encoding='utf-8') as file:
            for row in data:
                prompt = row[0]
                completion = row[1]
                JSON = {
                    "prompt": prompt + "\\n\\n###\\n\\n",
                    "completion": " " + completion + "\\n"
                }
                file.write(json.dumps(JSON) + '\n')
            return True
    except:
        return False

n = len(sys.argv)
if n!=2:
    print("Invalid number of arguments\nCorrect Command Schema : python '.\Completions Dataset Formatter.py' [Input Filename]")
else:
    input_file_path=sys.argv[1].replace('.\\','')
    output_file_path=input_file_path.split('.')[0] + '.jsonl'

    print("Input File Path : " + input_file_path)
    print("Output File Path : " + output_file_path)

    print("Check if input file exists Started")
    if not os.path.isfile(input_file_path):
        print("Check if input file exists Completed - Input file does not exist")
        exit()
    print("Check if input file exists Completed - Input file exists")

    if input_file_path.endswith('.xlsx'):
        print('Reading from XLSX file Started')
        data=read_from_excel(input_file_path)
        print('Reading from XLSX file Completed')
    elif input_file_path.endswith('.csv'):
        print('Reading from CSV file Started')
        data=read_from_csv(input_file_path)
        print('Reading from CSV file Completed')
    else:
        print("Only CSV and XLSX files are allowed as input")
        exit()
    
    print('Writing to JSONL Started')
    result=write_to_jsonl(data, output_file_path)
    if result is True:
        print('Writing to JSONL Completed')
    else:
        print('Writing to JSONL Failed')